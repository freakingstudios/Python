import openpyxl
from openpyxl.styles import Font, PatternFill

def search_and_highlight_text_in_all_sheets(input_file, output_file, texts_to_search):
    try:
        wb = openpyxl.load_workbook(input_file)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    for text in texts_to_search:
                        if text in str(cell.value):
                            cell.font = Font(bold=True)  # Make the text bold
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

        wb.save(output_file)
        print(f"Excel file '{input_file}' has been processed. Results saved to '{output_file}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # Replace 'input_file.xlsx' with the path to your input Excel file
    input_file_path = "input_file.xlsx"

    # Replace 'output_file.xlsx' with the desired output Excel file path
    output_file_path = "outputfile2.xlsx"

    # List of all the text strings to search
    texts_to_search = [
        "SCS1008726", "SCS1003245", "SCS1205765", "SCS1088041", "SCS1335778", "SCS1434217", "SCS1010751",
        "SCS1358363", "SCS1035215", "SCS1007581", "SCS1290767", "SCS1133963", "SCS1016802", "SCS1347806",
        "SCS1005544", "SCS1012323", "SCS1112284", "SCS1323238", "SCS1330036", "SCS1319008", "SCS1339102",
        "SCS1235275", "SCS1168283", "SCS1038411", "SCS1096455", "SCS1374160", "SCS1323361", "SCS1115085",
        "SCS935249", "SCS1106755", "SCS930004", "SCS941411", "SCS1447433", "SCS1190935", "SCS1461855",
        "SCS1012756", "SCS1030341", "SCS825643", "SCS860001", "SCS932248", "SCS1123904", "SCS1372386",
        "SCS817553", "SCS817545", "SCS1130884", "SCS891277", "SCS1041281", "SCS938669", "SCS941875",
        "SCS941879", "SCS1033341", "SCS961482", "SCS874951", "SCS804357", "SCS1328309", "SCS679367",
        "SCS1146733", "SCS1369144", "SCS1005496", "SCS1358359", "SCS1353819", "SCS1332801", "SCS679493",
        "SCS680565", "SCS681299", "SCS808362", "SCS856450", "SCS951826", "SCS835609", "SCS698209",
        "SCS804222", "SCS999334", "SCS823796", "SCS864314", "SCS1011274", "SCS695713", "SCS713777",
        "SCS573650", "SCS680541", "SCS682558", "SCS943252", "SCS628012", "SCS940816", "SCS821460",
        "SCS689031", "SCS674218", "SCS1013072", "SCS680440", "SCS693847", "SCS678969", "SCS676830",
        "SCS689911", "SCS679261", "SCS679054", "SCS685221", "SCS1164493", "SCS943511", "SCS697009",
        "SCS821234", "SCS639839", "SCS685189", "SCS1213303", "SCS1152936", "SCS1352994", "SCS955702",
        "SCS1027359", "SCS729925", "SCS1027842", "SCS1027843", "SCS698328", "SCS684284", "SCS576264",
        "SCS693626", "SCS590396", "SCS738203", "SCS589824", "SCS593806", "SCS581230", "SCS581227",
        "SCS1031236", "SCS961940", "SCS576345", "SCS577517", "SCS596428", "SCS858704", "SCS1125005",
        "SCS814669", "SCS681735", "SCS704907", "SCS1363258", "SCS592675", "SCS596735", "SCS1386023",
        "SCS1323703", "SCS1000424", "SCS1447426", "SCS578685", "SCS575165", "SCS1395588", "SCS1334762",
        "SCS616790", "SCS1474073", "SCS818414", "SCS574275", "SCS577131", "SCS677606", "SCS811179",
        "SCS596082", "SCS677517", "SCS891002", "SCS481300", "SCS619638", "SCS576347", "SCS1330050",
        "SCS1155968", "SCS839449", "SCS824242", "SCS1015358", "SCS1155358", "SCS812660", "SCS1427689",
        "SCS703393", "SCS598039", "SCS573641", "SCS575259", "SCS1469852", "SCS1016170", "SCS626038",
        "SCS1039804", "SCS468360", "SCS587523", "SCS1007792", "SCS692694", "SCS718213", "SCS734280",
        "SCS893566", "SCS587941", "SCS1364252", "SCS678062", "SCS951090", "SCS1105617", "SCS1351438",
        "SCS699015", "SCS679606", "SCS1327588", "SCS699565", "SCS1346980", "SCS723827", "SCS1195299",
        "SCS695097", "SCS935134", "SCS719293", "SCS586423", "SCS604037", "SCS1018137", "SCS679337",
        "SCS1334207", "SCS1324633", "SCS1427683", "SCS1391544", "SCS681008", "SCS1478472", "SCS575807",
        "SCS1410348", "SCS1015240", "SCS839503", "SCS1225855", "SCS1161181", "SCS1114876", "SCS935781",
        "SCS1009527", "SCS1346440", "SCS1180187", "SCS1150638", "SCS618144", "SCS944911"
    ]

    search_and_highlight_text_in_all_sheets(input_file_path, output_file_path, texts_to_search)
