import speech_recognition as sr
import openpyxl

# Initialize the recognizer class
r = sr.Recognizer()

# Get the audio from the microphone
with sr.Microphone() as source:
    print("Speak: ")
    audio = r.listen(source)

# Convert the audio to text
try:
    text = r.recognize_google(audio)
    print("You said: {}".format(text))
except sr.UnknownValueError:
    print("Sorry, I didn't catch that.")

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook.active

# Get the current number of rows in the sheet
rows = sheet.max_row + 1

# Write the text to the first column of the next row
sheet.cell(row=rows, column=1).value = text

# Save the changes to the workbook
workbook.save("data.xlsx")